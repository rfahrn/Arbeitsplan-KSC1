#!/usr/bin/env python3
"""
Arbeitskalender-Generator für KSC-Team (Version 5.0 - FINAL)
=============================================================
- Tagesverantwortung: GLEICHE Person für VM + NM am selben Tag
- Nur die 18 Kernmitarbeiter (ohne Stephi, Tamara, Manuela, Maria F., SPP)
- Legende exakt wie Referenzbild mit separater TV-Legende
- Wochenaufgaben: Je 1 Person für Direkt, ONB, BTM
"""

import json
import random
import os
from datetime import datetime, timedelta
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Tuple
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# KONFIGURATION
# ============================================================

DAYS = ["Montag", "Dienstag", "Mittwoch", "Donnerstag", "Freitag"]
SLOTS = ["Vormittag", "Nachmittag"]

# Farben für Aufgaben (EXAKT wie Referenzbild)
TASK_COLORS = {    # ===== Grün =====
    "KTG":        "A9D18E",
    "KGS":        "A9D18E",

    # ===== Orange =====
    "ABKL":       "FFC000",
    "PO/ABKL":    "FFC000",
    # ===== Telefon / Grün =====
    "TEL":        "92D050",
    "PO/TEL":     "92D050",
    "TEL-16:30":  "92D050",
    "TEL 16:00":  "92D050",

    # ===== ERF7 / Blau =====
    "ERF7":       "00B0F0",
    "ERF7/Q":     "00B0F0",
    "ERF7/HUB":   "00B0F0",

    # ===== ERF8 / Dunkelorange =====
    "ERF8":       "FF6600",
    "ERF8/Q":     "FF6600",
    "ERF8/HUB":   "FF6600",

    # ===== ERF9 / Rot =====
    "ERF9":       "FF0000",
    "ERF9/Q":     "FF0000",
    "ERF9/TEL":   "FF0000",

    # ===== Schalter / Hellgrün =====
    "ERF4/SCH":   "C6EFCE",
    "ERF7/SCH":   "C6EFCE",
    "ERF4/Access":"C6EFCE",

    # ===== Labor / Hellorange =====
    "ERF5":       "F4B084",

    # ===== Post / Scan / Grau =====
    "PO":         "D6DCE4",
    "PO/SCAN":    "D6DCE4",
    "SCAN":       "D6DCE4",

    # ===== RX Abo / Lila =====
    "RX Abo":     "D9B8FF",
    
    # ===== VBZ / Türkis =====
    "VBZ":        "4FD1C5",
    "VBZ/Q":      "4FD1C5",

    # ===== HO / Hellblau =====
    "HO":         "BDD7EE",
    "HO/Q":       "BDD7EE",


    # ===== Spezial =====
    "KSC Spez.":  "FFFF00",
    "TAGESPA":    "00FFFF",
    "KSV":        "00FFFF",

    # ===== Access / Magenta =====
    "Access":     "FF00FF",

    # ===== Abwesenheiten =====
    "Krankheit":  "FF9999",
    "KRANK":      "FF9999",
    "Feiertag":   "FF0000",
    "FERIEN":     "F2F2F2",

    # ===== Leer =====
    "":           "FFFFFF",
    
    
    #"TEL":       "92D050",  # Grün
    #"ERF7":      "00B0F0",  # Blau
    #"ERF7/Q":    "00B0F0",
    #"ERF7/HUB":  "00B0F0",
    # "ERF8":      "FF6600",  # Dunkelorange
    # "ERF8/Q":    "FF6600",
    # "ERF8/HUB":  "FF6600",

    # "ERF9":      "FF0000",  # Rot
    # "ERF9/Q":    "FF0000",
    # "ERF9/TEL":  "FF0000",

    # "ERF4/SCH":  "C6EFCE",  # Hellgrün - Schalter (legacy)
    # "ERF7/SCH":  "C6EFCE",  # Hellgrün - Schalter (kombiniert mit ERF7)

    # "ERF5":      "F4B084",  # Hellorange - Labor

    # "PO":        "D6DCE4",  # Grau - PÖ (Postöffnung) ohne Scan
    # "PO/ABKL":   "FFC000",
    # "PO/TEL":    "92D050",
    # "PO/SCAN":   "D6DCE4",  # Grau - PÖ + Scan (morgens)
    # "SCAN":      "D6DCE4",  # Grau - nur Scannen (nachmittags)

    # "HO":        "BDD7EE",  # Hellblau
    # "HO/Q":      "BDD7EE",

    # "VBZ/Q":     "C5E0B4",  # Lindgrün
    # "KSC Spez.": "FFFF00",  # Gelb
    # "TAGESPA":   "00FFFF",
    # "RX Abo":    "D9D9D9",  # Grau
    # "Access":    "FF00FF",  # Magenta
    # "KSV":       "00FFFF",
    # "Krankheit": "FF9999",
    # "KRANK":     "FF9999",
    # "Feiertag":  "FF0000",
    # "FERIEN":    "F2F2F2",
    # "TEL-16:30": "92D050",
    # "TEL 16:00": "92D050",
    # "ERF4/Access": "C6EFCE",
    # "":          "FFFFFF",
}

# Tagesverantwortung-Farben (für Header UND Legende)
TV_COLORS = {
    "Lara":      "00B050",  # Grün
    "Stephi":    "FF00FF",  # Pink
    "Silvana":   "FF6600",  # Orange
    "Linda":     "FFFF00",  # Gelb
}

# PHC-Liste = dunkelblau
PHC_COLOR = "0070C0"

# Berechtigungen für Wochenaufgaben
ONB_ERLAUBT = ["Silvana", "Linda", "Lara", "Andrea G.", "Dipiga",
               "Isaura", "Martina", "Alessia", "Amra", "Nina", "Jesika", "Corinne"]
DIREKT_ERLAUBT = ["Andrea A.", "Andrea G.", "Silvana"]
BTM_ERLAUBT = ["Jesika", "Silvana", "Linda", "Lara", "Dipiga", "Isaura"]

# ============================================================
# MITARBEITER (NUR die 18 Kernmitarbeiter!)
# ============================================================

@dataclass
class Employee:
    name: str
    short: str
    pct: int
    available: Dict[Tuple[int, int], bool] = field(default_factory=dict)
    can_hub: bool = False
    can_rx_abo: bool = False
    can_schalter: bool = True
    can_scanning: bool = True
    can_onb: bool = True
    can_direkt: bool = False
    can_vbz: bool = False
    is_tl: bool = False
    
    def is_available(self, day: int, slot: int) -> bool:
        return self.available.get((day, slot), False)


def create_employees(week_number: int) -> Dict[str, Employee]:
    """Erstellt NUR die 18 Kernmitarbeiter."""
    is_even_week = (week_number % 2 == 0)
    employees = {}
    
    def full_week(emp):
        for d in range(5):
            for s in range(2):
                emp.available[(d, s)] = True
    
    # Silvana - 100% (TL)
    e = Employee("Grossenbacher Silvana", "Silvana", 100, can_direkt=True, is_tl=True)
    full_week(e)
    employees["Silvana"] = e
    
    # Linda - 90%, jeden 2. Freitag frei (TL)
    e = Employee("Rexhaj Majlinda", "Linda", 90, can_rx_abo=True, is_tl=True)
    full_week(e)
    if is_even_week:
        e.available[(4, 0)] = False
        e.available[(4, 1)] = False
    employees["Linda"] = e
    
    # Lara - 100% (TL)
    e = Employee("Lara Ierano", "Lara", 100, can_rx_abo=True, can_vbz=True, can_hub=True, is_tl=True)
    full_week(e)
    employees["Lara"] = e
    
    # Andrea G. - 100%
    e = Employee("Gygax Andrea", "Andrea G.", 100, can_direkt=True)
    full_week(e)
    employees["Andrea G."] = e
    
    # Dipiga - 100%
    e = Employee("Jeyalingam Dipiga", "Dipiga", 100, can_hub=True)
    full_week(e)
    employees["Dipiga"] = e
    
    # Isaura - 90%, jeden 2. Mittwoch frei
    e = Employee("Bohnenblust Isaura", "Isaura", 90, can_rx_abo=True)
    full_week(e)
    if is_even_week:
        e.available[(2, 0)] = False
        e.available[(2, 1)] = False
    employees["Isaura"] = e
    
    # Martina - 80%, jeden Montag frei
    e = Employee("Martina Pizzi", "Martina", 80, can_rx_abo=True)
    full_week(e)
    e.available[(0, 0)] = False
    e.available[(0, 1)] = False
    employees["Martina"] = e
    
    # Alessia - 100%
    e = Employee("Giombanco Alessia", "Alessia", 100, can_hub=True)
    full_week(e)
    employees["Alessia"] = e
    
    # Amra - 100%
    e = Employee("Imsirovic Amra", "Amra", 100, can_hub=True)
    full_week(e)
    employees["Amra"] = e
    
    # Nina - 100%
    e = Employee("Hänni Nina", "Nina", 100, can_hub=True, can_vbz=True)
    full_week(e)
    employees["Nina"] = e
    
    # Jesika - 100%
    e = Employee("Bushaj Jesika", "Jesika", 100, can_hub=True)
    full_week(e)
    employees["Jesika"] = e
    
    # Brigitte - 70%, Mi/Do/Fr NM frei
    e = Employee("Siegrist Brigitte", "Brigitte", 70, can_schalter=False, can_scanning=False, can_onb=False)
    full_week(e)
    e.available[(2, 1)] = False  # Mi NM
    e.available[(3, 1)] = False  # Do NM
    e.available[(4, 1)] = False  # Fr NM
    employees["Brigitte"] = e
    
    # Florence - 70%, Mi NM + Fr frei
    e = Employee("Florence Dornbierer", "Florence", 70, can_schalter=False, can_onb=False)
    full_week(e)
    e.available[(2, 1)] = False  # Mi NM
    e.available[(4, 0)] = False  # Fr
    e.available[(4, 1)] = False
    employees["Florence"] = e
    
    # Corinne - 90%, jeden 2. Freitag frei
    e = Employee("Eggimann Corinne", "Corinne", 90, can_hub=True, can_schalter=False, can_scanning=False)
    full_week(e)
    if is_even_week:
        e.available[(4, 0)] = False
        e.available[(4, 1)] = False
    employees["Corinne"] = e
    
    # Saskia - 40%, nur Mi und Fr
    e = Employee("Schöni Saskia", "Saskia", 40, can_schalter=False, can_scanning=False, can_onb=False)
    for d in range(5):
        for s in range(2):
            e.available[(d, s)] = False
    e.available[(2, 0)] = True  # Mi
    e.available[(2, 1)] = True
    e.available[(4, 0)] = True  # Fr
    e.available[(4, 1)] = True
    employees["Saskia"] = e
    
    # Dragi - 60%, Di/Do/Fr hier
    e = Employee("Milenkovic Dragana", "Dragi", 60, can_hub=True, can_schalter=False, can_scanning=False, can_onb=False)
    for d in range(5):
        for s in range(2):
            e.available[(d, s)] = False
    e.available[(1, 0)] = True  # Di
    e.available[(1, 1)] = True
    e.available[(3, 0)] = True  # Do
    e.available[(3, 1)] = True
    e.available[(4, 0)] = True  # Fr
    e.available[(4, 1)] = True
    employees["Dragi"] = e
    
    # Maria B. - 30%, Mo und Do VM hier
    e = Employee("Bruzzese Maria", "Maria B.", 30, can_schalter=False, can_scanning=False, can_onb=False)
    for d in range(5):
        for s in range(2):
            e.available[(d, s)] = False
    e.available[(0, 0)] = True  # Mo VM
    e.available[(3, 0)] = True  # Do VM
    employees["Maria B."] = e
    
    # Andrea A. - 60%, Mo/Di/Mi hier
    e = Employee("Ackermann Andrea", "Andrea A.", 60, can_direkt=True, can_schalter=False, can_onb=False)
    for d in range(5):
        for s in range(2):
            e.available[(d, s)] = False
    e.available[(0, 0)] = True  # Mo
    e.available[(0, 1)] = True
    e.available[(1, 0)] = True  # Di
    e.available[(1, 1)] = True
    e.available[(2, 0)] = True  # Mi
    e.available[(2, 1)] = True
    employees["Andrea A."] = e
    
    return employees


# ============================================================
# SCHEDULE
# ============================================================

class Schedule:
    def __init__(self, employees, week_number, week_start_date):
        self.employees = employees
        self.week_number = week_number
        self.week_start_date = week_start_date
        self.is_even_week = (week_number % 2 == 0)
        
        self.schedule = {name: {} for name in employees}
        self.tel_count = {d: {s: 0 for s in range(2)} for d in range(5)}
        self.abkl_count = {d: {s: 0 for s in range(2)} for d in range(5)}
        self.erf8_assigned = {d: False for d in range(5)}
        
        # Tagesverantwortung: day -> name (GLEICHE Person für VM+NM!)
        self.tagesverantwortung: Dict[int, str] = {}
        
        # PHC-Liste: day -> name
        self.phc_liste: Dict[int, str] = {}
        
        # Wochenaufgaben: Je 1 Person pro Typ
        self.wochenaufgabe_direkt: str = ""
        self.wochenaufgabe_onb: str = ""
        self.wochenaufgabe_btm: str = ""
        
        # Notizen (z.B. Arzttermine) - (name, day, slot) -> note
        self.notes: Dict[Tuple[str, int, int], str] = {}
        
        self.conflicts = []
    
    def assign(self, name, day, slot, task):
        if name not in self.employees:
            return False
        if (day, slot) in self.schedule[name]:
            return False
        if not self.employees[name].is_available(day, slot):
            return False
        
        self.schedule[name][(day, slot)] = task
        if "TEL" in task:
            self.tel_count[day][slot] += 1
        if "ABKL" in task:
            self.abkl_count[day][slot] += 1
        if "ERF8" in task:
            self.erf8_assigned[day] = True
        return True
    
    def force_assign(self, name, day, slot, task):
        if name not in self.employees:
            return
        self.schedule[name][(day, slot)] = task
        if "TEL" in task:
            self.tel_count[day][slot] += 1
        if "ABKL" in task:
            self.abkl_count[day][slot] += 1
    
    def is_free(self, name, day, slot):
        return ((day, slot) not in self.schedule[name]
                and self.employees[name].is_available(day, slot))
    
    def get_task(self, name, day, slot):
        return self.schedule[name].get((day, slot), "")
    
    def get_available(self, day, slot, exclude=None):
        result = []
        for name, emp in self.employees.items():
            if exclude and name in exclude:
                continue
            if self.is_free(name, day, slot):
                result.append(name)
        return result


# ============================================================
# SCHEDULER
# ============================================================

def build_schedule(week_number, week_start_date, overrides=None, state_file="scheduler_state.json"):
    employees = create_employees(week_number)
    sched = Schedule(employees, week_number, week_start_date)
    
    state = load_state(state_file)
    
    last_week = state.get("last_generated_week", 0)
    is_new_week = (last_week != week_number)
    
    if is_new_week:
        state["last_generated_week"] = week_number
        state["linda_ho_week"] = not state.get("linda_ho_week", False)
        state["labor_jesika"] = not state.get("labor_jesika", True)
        state["schalter_group"] = 1 - state.get("schalter_group", 0)
        state["friday_phc_corinne"] = not state.get("friday_phc_corinne", True)
        state["tl_monday_idx"] = (state.get("tl_monday_idx", 0) + 1) % 3
        state["tv_start_day"] = (state.get("tv_start_day", 0) + 1) % 4
        state["onb_idx"] = (state.get("onb_idx", 0) + 1) % len(ONB_ERLAUBT)
        state["direkt_idx"] = (state.get("direkt_idx", 0) + 1) % len(DIREKT_ERLAUBT)
        state["btm_idx"] = (state.get("btm_idx", 0) + 1) % len(BTM_ERLAUBT)
    
    # Overrides verarbeiten
    if overrides:
        for (name, day, slot), task in overrides.items():
            if name in employees:
                if task.startswith("*"):
                    # Notiz (z.B. Arzttermin) - Person als abwesend markieren
                    sched.notes[(name, day, slot)] = task[1:]  # ohne *
                    sched.force_assign(name, day, slot, "")  # Slot blockieren
                elif task in ["KRANK", "FERIEN", "Krankheit"]:
                    sched.force_assign(name, day, slot, task)
                else:
                    sched.assign(name, day, slot, task)
    
    # ════════════════════════════════════════════════════════════
    # FIXE ZUWEISUNGEN
    # ════════════════════════════════════════════════════════════
    
    # Brigitte Mo ERF9 (ganztags)
    sched.assign("Brigitte", 0, 0, "ERF9")
    sched.assign("Brigitte", 0, 1, "ERF9")
    
    # Dragi Di ERF9 (ganztags)
    if employees["Dragi"].is_available(1, 0):
        sched.assign("Dragi", 1, 0, "ERF9")
    if employees["Dragi"].is_available(1, 1):
        sched.assign("Dragi", 1, 1, "ERF9/TEL")
    
    # Maria B. Mo HO
    if employees["Maria B."].is_available(0, 0):
        sched.assign("Maria B.", 0, 0, "HO")
    
    # Linda alle 2 Wochen Di HO
    if state.get("linda_ho_week", False):
        sched.assign("Linda", 1, 0, "HO/Q")
        sched.assign("Linda", 1, 1, "HO/Q")
    
    # Dipiga KGS: Di-morgens und Do-nachmittags
    sched.assign("Dipiga", 1, 0, "KGS")
    sched.assign("Dipiga", 3, 1, "KGS")
    
    # Martina KSC Spez. Di/Do VM
    if employees["Martina"].is_available(1, 0):
        sched.assign("Martina", 1, 0, "KSC Spez.")
    if employees["Martina"].is_available(3, 0):
        sched.assign("Martina", 3, 0, "KSC Spez.")
    
    # Labor: Jesika/Dipiga Mi NM
    if state.get("labor_jesika", True):
        sched.assign("Jesika", 2, 1, "ERF5")
    else:
        sched.assign("Dipiga", 2, 1, "ERF5")

    # ════════════════════════════════════════════════════════════
    # SCHALTER - 1 Person pro Tag, GANZER Tag (VM + NM dieselbe Person)
    # Zuerst zuweisen (vor TPA/RX Abo/HUB), damit ganztags-fähige
    # Personen reserviert werden. Task-Code: ERF7/SCH.
    # Rotation: schalter_group wechselt wöchentlich → 2-Wochen-Rhythmus.
    # ════════════════════════════════════════════════════════════

    schalter_exclude = {"Linda", "Florence", "Corinne", "Maria B.", "Andrea A.",
                        "Brigitte", "Saskia", "Dragi"}
    schalter_eligible = sorted([n for n in employees
                                if employees[n].can_schalter
                                and n not in schalter_exclude])

    schalter_group = state.get("schalter_group", 0)
    mid = len(schalter_eligible) // 2
    schalter_active = (schalter_eligible[:mid] if schalter_group == 0
                       else schalter_eligible[mid:])

    schalter_pool = list(schalter_active)
    random.shuffle(schalter_pool)
    schalter_used = set()
    for day in range(5):
        for name in schalter_pool:
            if name in schalter_used:
                continue
            if sched.is_free(name, day, 0) and sched.is_free(name, day, 1):
                sched.assign(name, day, 0, "ERF7/SCH")
                sched.assign(name, day, 1, "ERF7/SCH")
                schalter_used.add(name)
                break

    # ════════════════════════════════════════════════════════════
    # RX ABO - 2x pro Woche, halbtags, je 1 Person aus Pool
    # Pool: Linda, Lara, Isaura, Martina
    # ════════════════════════════════════════════════════════════

    rx_eligible = ["Linda", "Lara", "Isaura", "Martina"]
    random.shuffle(rx_eligible)
    rx_slots = [(d, s) for d in range(5) for s in range(2)]
    random.shuffle(rx_slots)

    rx_used_days = set()  # höchstens 1 RX Abo pro Tag
    rx_assigned = 0
    for (d, s) in rx_slots:
        if rx_assigned >= 2:
            break
        if d in rx_used_days:
            continue
        for name in rx_eligible:
            if sched.is_free(name, d, s):
                if sched.assign(name, d, s, "RX Abo"):
                    rx_used_days.add(d)
                    rx_assigned += 1
                    break

    # ════════════════════════════════════════════════════════════
    # TAGESPHARM (1 Person pro Tag, GANZER Tag = VM + NM)
    # Pro Person max. 1× pro Woche (keine Mehrfachzuweisung)
    # ════════════════════════════════════════════════════════════

    tpa_exclude = {"Brigitte", "Saskia", "Dragi", "Maria B.",
                   "Andrea A.", "Florence"}
    tpa_candidates = [n for n in employees
                      if n not in tpa_exclude
                      and not employees[n].is_tl]
    random.shuffle(tpa_candidates)
    tpa_used = set()

    for day in range(5):
        for name in tpa_candidates:
            if name in tpa_used:
                continue
            if sched.is_free(name, day, 0) and sched.is_free(name, day, 1):
                sched.assign(name, day, 0, "TAGES PA")
                sched.assign(name, day, 1, "TAGES PA")
                tpa_used.add(name)
                break

    # ════════════════════════════════════════════════════════════
    # ERF7/HUB - HALBTAGS, jeden Werktag 1 Person VM + 1 (andere) NM
    # ════════════════════════════════════════════════════════════

    hub_eligible = ["Jesika", "Dragi", "Lara", "Corinne", "Dipiga",
                    "Nina", "Amra", "Alessia"]

    for day in range(5):
        vm_pool = [n for n in hub_eligible if sched.is_free(n, day, 0)]
        random.shuffle(vm_pool)
        vm_pick = vm_pool[0] if vm_pool else None
        if vm_pick:
            sched.assign(vm_pick, day, 0, "ERF7/HUB")

        nm_pool = [n for n in hub_eligible
                   if n != vm_pick and sched.is_free(n, day, 1)]
        random.shuffle(nm_pool)
        nm_pick = nm_pool[0] if nm_pool else None
        if nm_pick:
            sched.assign(nm_pick, day, 1, "ERF7/HUB")

    # ════════════════════════════════════════════════════════════
    # PÖ (Postöffnung) + SCANNING - VOR TEL/ABKL einplanen, damit
    # die 3 PÖ-Plätze morgens garantiert sind.
    #   Morgens: 1 PO/SCAN + 2 PO  (3 PÖ-Personen)
    #   Nachmittag: 1 Scan (kein PÖ)
    # ════════════════════════════════════════════════════════════

    no_scanning = {"Brigitte", "Corinne", "Maria B.", "Saskia", "Dragi"}

    def _pick_one(day, slot, task, *, exclude_scanning):
        pool = [n for n in employees
                if (not exclude_scanning or n not in no_scanning)
                and not employees[n].is_tl
                and sched.is_free(n, day, slot)]
        random.shuffle(pool)
        for name in pool:
            if sched.assign(name, day, slot, task):
                return name
        return None

    def _pick_n_po(day, slot, count):
        pool = [n for n in employees
                if not employees[n].is_tl
                and sched.is_free(n, day, slot)]
        random.shuffle(pool)
        assigned = 0
        for name in pool:
            if assigned >= count:
                break
            if sched.assign(name, day, slot, "PO"):
                assigned += 1

    for day in range(5):
        # Vormittag: 1 PO/SCAN + 2 PO  → insgesamt 3 PÖ-Personen
        _pick_one(day, 0, "PO/SCAN", exclude_scanning=True)
        _pick_n_po(day, 0, 2)
        # Nachmittag: nur 1 Scan (kein PÖ)
        _pick_one(day, 1, "Scan", exclude_scanning=True)

    # ════════════════════════════════════════════════════════════
    # TEL (4 Mo VM, sonst 3)
    # ════════════════════════════════════════════════════════════
    
    tl_names = [n for n, emp in employees.items() if emp.is_tl]
    tl_tel_count = {n: 0 for n in tl_names}
    
    for day in range(5):
        for slot in range(2):
            target = 4 if (day == 0 and slot == 0) else 3
            needed = target - sched.tel_count[day][slot]
            if needed <= 0:
                continue
            
            available = sched.get_available(day, slot)
            tl_avail = [n for n in available if n in tl_tel_count and tl_tel_count[n] < 2]
            other = [n for n in available if n not in tl_names]
            random.shuffle(tl_avail)
            random.shuffle(other)
            
            count = 0
            for name in tl_avail + other:
                if count >= needed:
                    break
                if sched.assign(name, day, slot, "TEL"):
                    count += 1
                    if name in tl_tel_count:
                        tl_tel_count[name] += 1
    
    # ════════════════════════════════════════════════════════════
    # ABKL (2 pro Halbtag)
    # ════════════════════════════════════════════════════════════
    
    for day in range(5):
        for slot in range(2):
            needed = 2 - sched.abkl_count[day][slot]
            if needed <= 0:
                continue
            available = sched.get_available(day, slot)
            random.shuffle(available)
            count = 0
            for name in available:
                if count >= needed:
                    break
                if sched.assign(name, day, slot, "ABKL"):
                    count += 1
    
    # ════════════════════════════════════════════════════════════
    # ERF8 (1/Tag)
    # ════════════════════════════════════════════════════════════
    
    erf8_candidates = [n for n in employees if n not in ["Maria B.", "Saskia"]]
    random.shuffle(erf8_candidates)
    
    for day in range(5):
        if sched.erf8_assigned[day]:
            continue
        for name in erf8_candidates:
            if sched.is_free(name, day, 0):
                sched.assign(name, day, 0, "ERF8")
                break
            elif sched.is_free(name, day, 1):
                sched.assign(name, day, 1, "ERF8")
                break
    
    # ════════════════════════════════════════════════════════════
    # ERF9 Mi-Fr
    # ════════════════════════════════════════════════════════════
    
    erf9_candidates = [n for n in employees if n not in ["Brigitte", "Dragi"]]
    random.shuffle(erf9_candidates)
    
    for day in [2, 3, 4]:
        for name in erf9_candidates:
            if sched.is_free(name, day, 0):
                sched.assign(name, day, 0, "ERF9")
                break
    
    # ════════════════════════════════════════════════════════════
    # Restliche freie Slots auffüllen
    #   TLs            → ERF7/Q (Queue)
    #   alle anderen   → ERF7
    # WICHTIG: KEINE zusätzlichen "PO"-Zuweisungen mehr hier, damit
    # VM bei genau 1 PO/SCAN + 2 PO bleibt und nicht weitere
    # PÖ-Personen über die 2 hinaus aufgefüllt werden.
    # ════════════════════════════════════════════════════════════

    for name in employees:
        for day in range(5):
            for slot in range(2):
                if sched.is_free(name, day, slot):
                    if employees[name].is_tl:
                        sched.assign(name, day, slot, "ERF7/Q")
                    else:
                        sched.assign(name, day, slot, "ERF7")
    
    # ════════════════════════════════════════════════════════════
    # HILFSFUNKTION: Prüft ob Person an einem Tag verfügbar ist
    # ════════════════════════════════════════════════════════════
    
    def is_person_available_for_day(name, day):
        """Prüft ob Person für BEIDE Halbtage verfügbar ist (nicht KRANK/FERIEN/Termin)."""
        if name not in employees:
            return False
        # Prüfe VM
        task_vm = sched.get_task(name, day, 0)
        if task_vm in ["KRANK", "FERIEN", "Krankheit"] or task_vm.startswith("*"):
            return False
        if not employees[name].is_available(day, 0):
            return False
        # Prüfe NM
        task_nm = sched.get_task(name, day, 1)
        if task_nm in ["KRANK", "FERIEN", "Krankheit"] or task_nm.startswith("*"):
            return False
        if not employees[name].is_available(day, 1):
            return False
        return True
    
    def is_person_available_for_slot(name, day, slot):
        """Prüft ob Person für einen bestimmten Halbtag verfügbar ist."""
        if name not in employees:
            return False
        task = sched.get_task(name, day, slot)
        if task in ["KRANK", "FERIEN", "Krankheit"] or task.startswith("*"):
            return False
        return employees[name].is_available(day, slot)
    
    # ════════════════════════════════════════════════════════════
    # TAGESVERANTWORTUNG (GLEICHE Person für VM + NM!)
    # Prüft Verfügbarkeit und wählt Ersatz wenn nötig
    # ════════════════════════════════════════════════════════════
    
    tl_list = ["Lara", "Silvana", "Linda"]
    tv_start = state.get("tv_start_day", 0)
    
    for day in range(5):
        # Geplante TL für diesen Tag
        tl_idx = (tv_start + day) % 3
        planned_tl = tl_list[tl_idx]
        
        # Prüfe ob geplante TL verfügbar ist
        if is_person_available_for_day(planned_tl, day):
            sched.tagesverantwortung[day] = planned_tl
        else:
            # Suche Ersatz unter den anderen TLs
            replacement_found = False
            for i in range(1, 3):  # Probiere die anderen 2 TLs
                alt_idx = (tl_idx + i) % 3
                alt_tl = tl_list[alt_idx]
                if is_person_available_for_day(alt_tl, day):
                    sched.tagesverantwortung[day] = alt_tl
                    sched.conflicts.append(
                        f"TV {DAYS[day]}: {planned_tl} nicht verfügbar → {alt_tl} übernimmt"
                    )
                    replacement_found = True
                    break
            
            if not replacement_found:
                # Keine TL verfügbar - kritischer Konflikt
                sched.tagesverantwortung[day] = f"⚠️ OFFEN"
                sched.conflicts.append(
                    f"❌ TV {DAYS[day]}: Keine TL verfügbar! Manuell zuweisen."
                )
    
    # ════════════════════════════════════════════════════════════
    # PHC-LISTE (prüft ebenfalls Verfügbarkeit)
    # ════════════════════════════════════════════════════════════
    
    # Montag: TL im Wechsel (prüfe Verfügbarkeit)
    tl_monday = ["Silvana", "Linda", "Lara"]
    tl_idx = state.get("tl_monday_idx", 0)
    planned_phc_mo = tl_monday[tl_idx % 3]
    
    if is_person_available_for_slot(planned_phc_mo, 0, 1):  # NM muss da sein
        sched.phc_liste[0] = planned_phc_mo
    else:
        # Suche Ersatz
        for i in range(1, 3):
            alt_tl = tl_monday[(tl_idx + i) % 3]
            if is_person_available_for_slot(alt_tl, 0, 1):
                sched.phc_liste[0] = alt_tl
                sched.conflicts.append(f"PHC Mo: {planned_phc_mo} nicht da → {alt_tl}")
                break
        else:
            sched.phc_liste[0] = "⚠️ OFFEN"
            sched.conflicts.append("❌ PHC Mo: Keine TL verfügbar!")
    
    # Dienstag: Dipiga (prüfe Verfügbarkeit)
    if is_person_available_for_slot("Dipiga", 1, 1):
        sched.phc_liste[1] = "Dipiga"
    else:
        # Suche Ersatz
        available_di = [n for n in employees 
                       if is_person_available_for_slot(n, 1, 1) 
                       and n not in ["Saskia", "Maria B."]]
        if available_di:
            replacement = random.choice(available_di)
            sched.phc_liste[1] = replacement
            sched.conflicts.append(f"PHC Di: Dipiga nicht da → {replacement}")
        else:
            sched.phc_liste[1] = "⚠️ OFFEN"
            sched.conflicts.append("❌ PHC Di: Niemand verfügbar!")
    
    # Mittwoch: undefiniert, wähle aus verfügbaren
    available_mi = [n for n in employees 
                   if is_person_available_for_slot(n, 2, 1)
                   and n not in ["Saskia", "Maria B."]]
    sched.phc_liste[2] = random.choice(available_mi) if available_mi else "⚠️ OFFEN"
    
    # Donnerstag: undefiniert, wähle aus verfügbaren
    available_do = [n for n in employees 
                   if is_person_available_for_slot(n, 3, 1)
                   and n not in ["Saskia", "Maria B."]]
    sched.phc_liste[3] = random.choice(available_do) if available_do else "⚠️ OFFEN"
    
    # Freitag: Corinne/Andrea G. im Wechsel (prüfe Verfügbarkeit)
    planned_phc_fr = "Corinne" if state.get("friday_phc_corinne", True) else "Andrea G."
    alt_phc_fr = "Andrea G." if planned_phc_fr == "Corinne" else "Corinne"
    
    if is_person_available_for_slot(planned_phc_fr, 4, 1):
        sched.phc_liste[4] = planned_phc_fr
    elif is_person_available_for_slot(alt_phc_fr, 4, 1):
        sched.phc_liste[4] = alt_phc_fr
        sched.conflicts.append(f"PHC Fr: {planned_phc_fr} nicht da → {alt_phc_fr}")
    else:
        # Suche jemand anderen
        available_fr = [n for n in employees 
                       if is_person_available_for_slot(n, 4, 1)
                       and n not in ["Saskia", "Maria B."]]
        if available_fr:
            replacement = random.choice(available_fr)
            sched.phc_liste[4] = replacement
            sched.conflicts.append(f"PHC Fr: Weder {planned_phc_fr} noch {alt_phc_fr} da → {replacement}")
        else:
            sched.phc_liste[4] = "⚠️ OFFEN"
            sched.conflicts.append("❌ PHC Fr: Niemand verfügbar!")
    
    # ════════════════════════════════════════════════════════════
    # WOCHENAUFGABEN (je 1 Person pro Typ)
    # ════════════════════════════════════════════════════════════
    
    sched.wochenaufgabe_direkt = DIREKT_ERLAUBT[state.get("direkt_idx", 0) % len(DIREKT_ERLAUBT)]
    sched.wochenaufgabe_onb = ONB_ERLAUBT[state.get("onb_idx", 0) % len(ONB_ERLAUBT)]
    sched.wochenaufgabe_btm = BTM_ERLAUBT[state.get("btm_idx", 0) % len(BTM_ERLAUBT)]
    
    save_state(state, state_file)
    return sched


def load_state(filepath):
    if os.path.exists(filepath):
        with open(filepath, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {}


def save_state(state, filepath):
    with open(filepath, 'w', encoding='utf-8') as f:
        json.dump(state, f, indent=2)


# ============================================================
# EXCEL AUSGABE
# ============================================================

def write_excel(sched, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Arbeitsplan"
    
    start = sched.week_start_date
    end = start + timedelta(days=4)
    date_str = f"{start.strftime('%d.%m.%Y')}-{end.strftime('%d.%m.%Y')}"
    
    thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'), bottom=Side(style='thin'))
    hdr_font = Font(bold=True, size=10, name='Arial')
    cell_font = Font(size=9, name='Arial')
    leg_font = Font(size=8, name='Arial')
    
    # ════════════════════════════════════════════════════════════
    # LEGENDE (Zeilen 1-6)
    # ════════════════════════════════════════════════════════════
    
    # Zeile 1
    ws['A1'] = "Legende"
    ws['A1'].font = hdr_font
    
    legend1 = [("KTG", "A9D18E"), ("ABKL", "FFC000"), ("TEL", "92D050"),
               ("", "FFFFFF"), ("ERF7", "00B0F0"), ("PO", "D6DCE4"),
               ("PHC-Liste", "0070C0"), ("ERF9", "FF0000"), ("Acces", "FF00FF")]
    for i, (txt, col) in enumerate(legend1):
        c = ws.cell(row=1, column=3+i, value=txt)
        c.font = leg_font
        c.alignment = Alignment(horizontal='center')
        if col:
            c.fill = PatternFill('solid', fgColor=col)
            if col in ["0070C0", "FF0000"]:
                c.font = Font(size=8, color='FFFFFF')
    
    # Wochenaufgabe Header
    ws.cell(row=1, column=13, value="Wochen").font = leg_font
    ws.cell(row=1, column=14, value="Aufgabe").font = leg_font
    ws.cell(row=1, column=15, value="Bemerkungen").font = hdr_font
    
    # TV-Legende (rechts)
    ws.cell(row=2, column=13, value="Lara").fill = PatternFill('solid', fgColor='00B050')
    ws.cell(row=2, column=13).font = Font(size=8, color='FFFFFF')
    ws.cell(row=2, column=14, value="Stephi").fill = PatternFill('solid', fgColor='FF00FF')
    ws.cell(row=2, column=14).font = Font(size=8, color='FFFFFF')
    
    # Zeile 2: Erklärung
    ws['A2'] = "Erklärung"
    ws['B2'] = "AE"
    ws['B2'].fill = PatternFill('solid', fgColor='FFC000')
    expl1 = ["Kein Tages-geschäft", "Abklärung", "Telefon", "Frei", 
             "Erfassung", "Post-öffnung", "PHC-Liste", "17.30 Dienst", ""]
    for i, txt in enumerate(expl1):
        ws.cell(row=2, column=3+i, value=txt).font = leg_font
    
    # Zeile 3
    ws['A3'] = "Legende"
    legend3 = [("HO", "BDD7EE"), ("Krankheit", "FF9999"), ("scanning", "E2EFDA"),
               ("PO/SCAN", "A9D18E"), ("KSV", "00FFFF"), ("ERF4/SCH", "C6EFCE"),
               ("ERF5", "F4B084"), ("Feiertag", "FF0000")]
    for i, (txt, col) in enumerate(legend3):
        c = ws.cell(row=3, column=3+i, value=txt)
        c.font = leg_font
        if col:
            c.fill = PatternFill('solid', fgColor=col)
    
    # TV-Legende Fortsetzung
    ws.cell(row=3, column=12, value="Silvana").fill = PatternFill('solid', fgColor='FF6600')
    ws.cell(row=3, column=12).font = Font(size=8, color='FFFFFF')
    ws.cell(row=4, column=12, value="Linda").fill = PatternFill('solid', fgColor='FFFF00')
    ws.cell(row=4, column=12).font = Font(size=8)
    
    # Zeile 4: Erklärung
    ws['A4'] = "Erklärung"
    expl2 = ["Home Office", "Krankheit", "Scanning / Ablage", "Postöffnung & Scanning",
             "KSV Aushilfe", "Schalter", "Labor", ""]
    for i, txt in enumerate(expl2):
        ws.cell(row=4, column=3+i, value=txt).font = leg_font
    
    # Zeile 5-6: KD
    ws['A5'] = "Legende"
    ws['B5'] = "KD"
    ws['B5'].fill = PatternFill('solid', fgColor='E2EFDA')
    legend5 = [("ID", "BDD7EE"), ("Krankheit", "FF9999"), ("scanning", "E2EFDA"),
               ("PO/SCAN", "A9D18E"), ("", "FFFFFF"), ("KD KOMM", "FFFFFF"),
               ("AE", "FFC000"), ("", "FFFFFF"), ("ERP", "C6EFCE"), ("Cafeteria", "FF0000")]
    for i, (txt, col) in enumerate(legend5):
        c = ws.cell(row=5, column=3+i, value=txt)
        c.font = leg_font
        if col:
            c.fill = PatternFill('solid', fgColor=col)
    
    ws['A6'] = "Erklärung"
    expl3 = ["Identifikation", "Krank", "Scanning / Ablage", "Postöffnung & Scanning",
             "Frei", "Postfach KD KOMM", "im AP AE schauen", "Spätdienst bis 17.30",
             "ERP Testen", ""]
    for i, txt in enumerate(expl3):
        ws.cell(row=6, column=3+i, value=txt).font = leg_font
    
    # ════════════════════════════════════════════════════════════
    # DATUM (Zeile 8)
    # ════════════════════════════════════════════════════════════
    
    ws.merge_cells('A8:L8')
    ws['A8'] = date_str
    ws['A8'].font = Font(bold=True, size=12, name='Arial')
    ws['A8'].alignment = Alignment(horizontal='center')
    
    # ════════════════════════════════════════════════════════════
    # HEADER (Zeile 10-11)
    # ════════════════════════════════════════════════════════════
    
    row = 10
    headers = ["Name", "%"]
    for day in DAYS:
        headers.extend([day, ""])
    headers.extend(["Wochen Aufgabe", "Bemerkungen"])
    
    for col, val in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(bold=True, size=10, color='FFFFFF')
        c.alignment = Alignment(horizontal='center')
        c.fill = PatternFill('solid', fgColor='4472C4')
        c.border = thin
    
    for i in range(5):
        start_col = 3 + i * 2
        ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col+1)
    
    # Zeile 11: VM/NM mit Tagesverantwortungs-Farben (GLEICHE Farbe für VM+NM!)
    row = 11
    ws.cell(row=row, column=1, value="").border = thin
    ws.cell(row=row, column=2, value="").border = thin
    
    for day in range(5):
        tv_name = sched.tagesverantwortung.get(day, "")
        tv_color = TV_COLORS.get(tv_name, "D9E2F3")
        
        for slot in range(2):
            col = 3 + day * 2 + slot
            c = ws.cell(row=row, column=col, value=SLOTS[slot])
            c.font = Font(bold=True, size=8)
            c.alignment = Alignment(horizontal='center')
            c.border = thin
            c.fill = PatternFill('solid', fgColor=tv_color)
            if tv_name in ["Lara", "Silvana"]:
                c.font = Font(bold=True, size=8, color='FFFFFF')
    
    ws.cell(row=row, column=13, value="").border = thin
    ws.cell(row=row, column=14, value="").border = thin
    
    # ════════════════════════════════════════════════════════════
    # MITARBEITER
    # ════════════════════════════════════════════════════════════
    
    order = ["Silvana", "Linda", "Lara", "Andrea G.", "Dipiga",
             "Isaura", "Martina", "Alessia", "Amra", "Nina",
             "Jesika", "Brigitte", "Florence", "Corinne", "Saskia",
             "Dragi", "Maria B.", "Andrea A."]
    
    row = 12
    for name in order:
        if name not in sched.employees:
            continue
        emp = sched.employees[name]
        
        ws.cell(row=row, column=1, value=emp.name).font = cell_font
        ws.cell(row=row, column=1).border = thin
        
        ws.cell(row=row, column=2, value=emp.pct).font = cell_font
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=2).border = thin
        
        for day in range(5):
            for slot in range(2):
                col = 3 + day * 2 + slot
                task = sched.get_task(name, day, slot)
                
                c = ws.cell(row=row, column=col, value=task)
                c.font = cell_font
                c.alignment = Alignment(horizontal='center')
                c.border = thin
                
                color = get_task_color(task)
                if color:
                    c.fill = PatternFill('solid', fgColor=color)
                
                # PHC blau
                if slot == 1 and sched.phc_liste.get(day) == name:
                    c.fill = PatternFill('solid', fgColor=PHC_COLOR)
                    c.font = Font(bold=True, size=9, color='FFFFFF')
        
        # Wochenaufgabe
        wa_cell = ws.cell(row=row, column=13)
        wa_cell.border = thin
        wa_parts = []
        if name == sched.wochenaufgabe_direkt:
            wa_parts.append("Direkt")
        if name == sched.wochenaufgabe_onb:
            wa_parts.append("ONB")
        if name == sched.wochenaufgabe_btm:
            wa_parts.append("BTM")
        if wa_parts:
            wa_cell.value = "\n".join(wa_parts)
            wa_cell.font = Font(bold=True, size=9)
            wa_cell.alignment = Alignment(horizontal='center', wrap_text=True)
        
        # Bemerkungen (inkl. Termine/Notizen)
        bem_parts = []
        for day in range(5):
            for slot in range(2):
                note_key = (name, day, slot)
                if note_key in sched.notes:
                    bem_parts.append(f"{DAYS[day][:2]} {SLOTS[slot][:2]}: {sched.notes[note_key]}")
        
        bem_cell = ws.cell(row=row, column=14)
        bem_cell.border = thin
        if bem_parts:
            bem_cell.value = "\n".join(bem_parts)
            bem_cell.font = cell_font
            bem_cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        row += 1
    
    # Spaltenbreiten
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 5
    for i in range(5):
        ws.column_dimensions[get_column_letter(3 + i * 2)].width = 12
        ws.column_dimensions[get_column_letter(4 + i * 2)].width = 12
    ws.column_dimensions[get_column_letter(13)].width = 12
    ws.column_dimensions[get_column_letter(14)].width = 15
    
    wb.save(output_path)
    return output_path


def get_task_color(task):
    if not task:
        return None
    if task in TASK_COLORS:
        return TASK_COLORS[task]
    for key in sorted(TASK_COLORS.keys(), key=len, reverse=True):
        if task.startswith(key):
            return TASK_COLORS[key]
    return None


def validate_schedule(sched):
    issues = []
    
    # Konflikte aus dem Scheduling-Prozess
    if sched.conflicts:
        for conflict in sched.conflicts:
            issues.append(conflict)
    
    # TEL Mindestbesetzung
    if sched.tel_count[0][0] < 4:
        issues.append(f"❌ Mo VM TEL: {sched.tel_count[0][0]}/4")
    for d in range(5):
        for s in range(2):
            if d == 0 and s == 0:
                continue
            if sched.tel_count[d][s] < 3:
                issues.append(f"❌ {DAYS[d]} {SLOTS[s]} TEL: {sched.tel_count[d][s]}/3")
    
    # ABKL Mindestbesetzung
    for d in range(5):
        for s in range(2):
            if sched.abkl_count[d][s] < 2:
                issues.append(f"❌ {DAYS[d]} {SLOTS[s]} ABKL: {sched.abkl_count[d][s]}/2")
    
    # ERF8 pro Tag
    for d in range(5):
        if not sched.erf8_assigned[d]:
            issues.append(f"❌ {DAYS[d]} ERF8: Nicht besetzt")
    
    return issues


def get_next_monday(ref=None):
    if ref is None:
        ref = datetime.now()
    days = 0 - ref.weekday()
    if days <= 0:
        days += 7
    return ref + timedelta(days=days)


# ============================================================
# INTERAKTIVE EINGABE VON SONDERWÜNSCHEN
# ============================================================

EMPLOYEE_LIST = [
    "Silvana", "Linda", "Lara", "Andrea G.", "Dipiga",
    "Isaura", "Martina", "Alessia", "Amra", "Nina",
    "Jesika", "Brigitte", "Florence", "Corinne", "Saskia",
    "Dragi", "Maria B.", "Andrea A."
]

def print_employees():
    print("\n  Mitarbeiterinnen:")
    for i, name in enumerate(EMPLOYEE_LIST, 1):
        print(f"    {i:2d}. {name}")

def pick_employee():
    """Lässt den User eine Mitarbeiterin wählen."""
    print_employees()
    while True:
        inp = input("\n  Nummer oder Name (Enter = fertig): ").strip()
        if not inp:
            return None
        if inp.isdigit():
            idx = int(inp) - 1
            if 0 <= idx < len(EMPLOYEE_LIST):
                return EMPLOYEE_LIST[idx]
        matches = [n for n in EMPLOYEE_LIST if inp.lower() in n.lower()]
        if len(matches) == 1:
            return matches[0]
        elif len(matches) > 1:
            print(f"    Mehrere Treffer: {', '.join(matches)}")
        else:
            print(f"    '{inp}' nicht gefunden.")

def pick_days():
    """Lässt den User Tage wählen."""
    print(f"\n  Welche Tage?")
    print(f"    1=Mo  2=Di  3=Mi  4=Do  5=Fr")
    print(f"    Mehrere mit Komma (z.B. '1,3,5') oder 'alle'")
    while True:
        inp = input("  Tage: ").strip().lower()
        if inp in ["alle", "all", "a"]:
            return [0, 1, 2, 3, 4]
        try:
            days = [int(x.strip()) - 1 for x in inp.split(",")]
            if all(0 <= d <= 4 for d in days):
                return days
        except ValueError:
            pass
        print("    Ungültige Eingabe. Beispiel: '1,3' für Mo+Mi")

def pick_slots():
    """Lässt den User Halbtage wählen."""
    print(f"\n  Welche Halbtage?")
    print(f"    1=Vormittag  2=Nachmittag  3=Ganzer Tag")
    while True:
        inp = input("  Halbtag: ").strip()
        if inp == "1":
            return [0]
        elif inp == "2":
            return [1]
        elif inp in ["3", "beide", "ganzer tag", "ganz"]:
            return [0, 1]
        print("    Bitte 1, 2 oder 3 eingeben.")

def pick_absence_type():
    """Lässt den User den Abwesenheitstyp wählen."""
    print(f"\n  Was ist der Grund?")
    print(f"    1 = Krank (ganzer Tag/Tage)")
    print(f"    2 = Ferien / Abwesend (ganzer Tag/Tage)")
    print(f"    3 = Arzttermin / fixer Termin (Halbtag)")
    print(f"    4 = Anderer Sonderwunsch (z.B. HO, bestimmte Aufgabe)")
    while True:
        inp = input("  Typ (1-4): ").strip()
        if inp in ["1", "2", "3", "4"]:
            return int(inp)
        print("    Bitte 1-4 eingeben.")

def check_rule_impact(absences, employees):
    """Prüft welche Regeln durch Abwesenheiten verletzt werden könnten."""
    warnings = []
    
    for name, day, slot in absences:
        day_name = DAYS[day]
        slot_name = SLOTS[slot]
        tag = f"{name} {day_name} {slot_name}"
        
        # Brigitte Mo ERF9
        if name == "Brigitte" and day == 0:
            warnings.append(f"⚠️  {tag}: Brigitte hat normalerweise ERF9 am Montag!")
        
        # Dragi Di ERF9
        if name == "Dragi" and day == 1:
            warnings.append(f"⚠️  {tag}: Dragi hat normalerweise ERF9 am Dienstag!")
        
        # Dipiga TAGES PA / KGS
        if name == "Dipiga" and day == 1 and slot == 0:
            warnings.append(f"⚠️  {tag}: Dipiga hat TAGES PA am Dienstag VM!")
        if name == "Dipiga" and day == 3 and slot == 1:
            warnings.append(f"⚠️  {tag}: Dipiga hat KGS am Donnerstag NM!")
        
        # Martina KSC Spez.
        if name == "Martina" and day in [1, 3] and slot == 0:
            warnings.append(f"⚠️  {tag}: Martina hat KSC Spez. am {day_name} VM!")
        
        # Florence PO Di
        if name == "Florence" and day == 1:
            warnings.append(f"⚠️  {tag}: Florence hat normalerweise PO am Dienstag!")
        
        # Maria B. HO Mo
        if name == "Maria B." and day == 0 and slot == 0:
            warnings.append(f"⚠️  {tag}: Maria B. hat normalerweise HO am Montag VM!")
        
        # ERF5 Mi NM
        if name in ["Jesika", "Dipiga"] and day == 2 and slot == 1:
            warnings.append(f"⚠️  {tag}: {name} könnte ERF5 (Labor) am Mi NM haben!")
    
    # Prüfe Mindestbesetzung
    for day in range(5):
        for slot in range(2):
            absent_count = sum(1 for n, d, s in absences if d == day and s == slot)
            available_count = sum(1 for n in employees 
                                 if employees[n].is_available(day, slot)) - absent_count
            target_tel = 4 if (day == 0 and slot == 0) else 3
            min_needed = target_tel + 2  # TEL + ABKL
            if available_count < min_needed:
                warnings.append(
                    f"🔴 {DAYS[day]} {SLOTS[slot]}: Nur {available_count} Personen verfügbar! "
                    f"Brauche mind. {min_needed} (TEL:{target_tel} + ABKL:2)!"
                )
            elif available_count < min_needed + 3:
                warnings.append(
                    f"🟡 {DAYS[day]} {SLOTS[slot]}: Nur {available_count} Personen verfügbar - eng!"
                )
    
    return warnings

def interactive_input(week_number):
    """Interaktive Eingabe von Abwesenheiten und Sonderwünschen."""
    employees = create_employees(week_number)
    overrides = {}
    absences = []
    
    print("\n" + "─" * 60)
    print("  SONDERWÜNSCHE / ABWESENHEITEN EINGEBEN")
    print("─" * 60)
    print("  Hier kannst du Krankmeldungen, Termine, Ferien etc. eingeben.")
    print("  Am Ende prüft das Programm welche Regeln betroffen sind.")
    
    entry_count = 0
    
    while True:
        print("\n" + "─" * 40)
        if entry_count == 0:
            prompt = "  Gibt es eine Abwesenheit oder einen Sonderwunsch? (j/n): "
        else:
            prompt = "  Noch eine Abwesenheit oder ein Sonderwunsch? (j/n): "
        ans = input(prompt).strip().lower()
        if ans not in ["j", "ja", "y", "yes"]:
            break
        
        name = pick_employee()
        if not name:
            break
        
        typ = pick_absence_type()
        
        if typ in [1, 2]:
            label = "KRANK" if typ == 1 else "FERIEN"
            days = pick_days()
            for day in days:
                for slot in [0, 1]:
                    if employees[name].is_available(day, slot):
                        overrides[(name, day, slot)] = label
                        absences.append((name, day, slot))
            day_str = ", ".join(DAYS[d] for d in days)
            print(f"\n  ✓ {name} → {label} am {day_str}")
            entry_count += 1
        
        elif typ == 3:
            days = pick_days()
            slots = pick_slots()
            notiz = input("  Notiz (z.B. 'Arzttermin 14:30', Enter=leer): ").strip()
            if not notiz:
                notiz = "Termin"
            for day in days:
                for slot in slots:
                    if employees[name].is_available(day, slot):
                        # Termin = Person abwesend für diesen Slot
                        overrides[(name, day, slot)] = f"*{notiz}"
                        absences.append((name, day, slot))
            day_str = ", ".join(DAYS[d] for d in days)
            slot_str = "+".join(SLOTS[s] for s in slots)
            print(f"\n  ✓ {name} → {notiz} am {day_str} ({slot_str})")
            entry_count += 1
        
        elif typ == 4:
            days = pick_days()
            slots = pick_slots()
            print(f"\n  Welche Aufgabe? (z.B. HO, TEL, ERF7, ABKL, etc.)")
            task = input("  Aufgabe: ").strip()
            if not task:
                task = "HO"
            for day in days:
                for slot in slots:
                    if employees[name].is_available(day, slot):
                        overrides[(name, day, slot)] = task
            day_str = ", ".join(DAYS[d] for d in days)
            slot_str = "+".join(SLOTS[s] for s in slots)
            print(f"\n  ✓ {name} → {task} am {day_str} ({slot_str})")
            entry_count += 1
    
    # Zusammenfassung
    if overrides:
        print("\n" + "═" * 60)
        print("  EINGEGEBENE SONDERWÜNSCHE:")
        print("═" * 60)
        for (name, day, slot), task in overrides.items():
            print(f"  • {name:15s}  {DAYS[day]:12s} {SLOTS[slot]:12s}  →  {task}")
        
        if absences:
            warnings = check_rule_impact(absences, employees)
            if warnings:
                print("\n" + "═" * 60)
                print("  AUSWIRKUNGEN AUF REGELN:")
                print("═" * 60)
                for w in warnings:
                    print(f"  {w}")
                print()
                ans = input("  Trotzdem weiterfahren? (j/n): ").strip().lower()
                if ans not in ["j", "ja", "y", "yes"]:
                    print("  Abgebrochen.")
                    return None
            else:
                print("\n  ✅ Keine kritischen Regel-Auswirkungen erkannt.")
    else:
        print("\n  Keine Sonderwünsche → normaler Plan wird generiert.")
    
    return overrides


def main():
    print()
    print("╔══════════════════════════════════════════════════════════╗")
    print("║         ARBEITSKALENDER-GENERATOR v5.1                  ║")
    print("║         Mit interaktiver Sonderwunsch-Eingabe           ║")
    print("╚══════════════════════════════════════════════════════════╝")
    
    monday = get_next_monday()
    kw = monday.isocalendar()[1]
    friday = monday + timedelta(days=4)
    
    print(f"\n  Woche:  KW {kw}")
    print(f"  Von:    {monday.strftime('%d.%m.%Y')} (Montag)")
    print(f"  Bis:    {friday.strftime('%d.%m.%Y')} (Freitag)")
    
    # 2-Wochen Info anzeigen
    is_even = kw % 2 == 0
    print(f"\n  2-Wochen-Status (KW {kw} = {'gerade' if is_even else 'ungerade'}):")
    if is_even:
        print("    • Linda: Freitag FREI")
        print("    • Isaura: Mittwoch FREI")
        print("    • Corinne: Freitag FREI")
    else:
        print("    • Linda, Isaura, Corinne: normal verfügbar")
    
    # Interaktive Eingabe
    overrides = interactive_input(kw)
    if overrides is None:
        return None
    
    print("\n  Generiere Arbeitsplan...")
    random.seed(kw)
    
    sched = build_schedule(kw, monday, overrides)
    issues = validate_schedule(sched)
    
    filename = f"Arbeitsplan_KW{kw}_{monday.strftime('%Y%m%d')}.xlsx"
    write_excel(sched, filename)
    
    # Ergebnis
    print(f"\n{'═' * 60}")
    print(f"  ERGEBNIS")
    print(f"{'═' * 60}")
    
    if issues:
        print(f"\n  ⚠️ ACHTUNG - Folgende Regeln konnten nicht eingehalten werden:")
        for issue in issues:
            print(f"    {issue}")
    else:
        print(f"\n  ✅ Alle Regeln eingehalten!")
    
    print(f"\n  📄 Datei erstellt: {filename}")
    
    # Zusammenfassung
    print(f"\n{'─' * 60}")
    print(f"  BESETZUNG:")
    print(f"{'─' * 60}")
    for day in range(5):
        for slot in range(2):
            tel_c = sched.tel_count[day][slot]
            abkl_c = sched.abkl_count[day][slot]
            target_tel = 4 if (day == 0 and slot == 0) else 3
            tel_sym = "✅" if tel_c >= target_tel else "❌"
            abkl_sym = "✅" if abkl_c >= 2 else "❌"
            print(f"  {DAYS[day]:12s} {SLOTS[slot]:12s}  "
                  f"TEL: {tel_c}/{target_tel} {tel_sym}  "
                  f"ABKL: {abkl_c}/2 {abkl_sym}")
    
    print(f"\n  Wochenaufgaben:")
    print(f"    Direkt: {sched.wochenaufgabe_direkt}")
    print(f"    ONB:    {sched.wochenaufgabe_onb}")
    print(f"    BTM:    {sched.wochenaufgabe_btm}")
    
    print(f"\n  Tagesverantwortung:")
    for d in range(5):
        print(f"    {DAYS[d]}: {sched.tagesverantwortung.get(d, '?')}")
    
    print(f"\n{'═' * 60}")
    print(f"  Fertig!")
    print(f"{'═' * 60}\n")
    
    return filename


if __name__ == "__main__":
    main()
